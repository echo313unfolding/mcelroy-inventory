#!/usr/bin/env python3
"""
Build a fillable Excel workbook from the McElroy parts database.
One file that serves two purposes:
  1. Email to shop manager to fill in on-hand counts, serial numbers, locations
  2. Import back into InvenTree (Parts → Import) after it's filled in

Output: UGSI_McElroy_Parts_Inventory.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from parts_data import CATEGORIES, PARTS, STOCK_LOCATIONS, MACHINE_MODELS, FLEET_TEMPLATE

# ── Style constants ──────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")  # McElroy green
SECTION_FILL = PatternFill(start_color="D9E2D0", end_color="D9E2D0", fill_type="solid")  # light green
SERIAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # yellow = fill me
INPUT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")    # blue = fill me
LOCKED_FONT = Font(name="Calibri", size=10, color="666666")
INPUT_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5233")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

LOCATIONS = [loc["name"] for loc in STOCK_LOCATIONS]


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_input=False, is_serial=False):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if is_serial:
        cell.fill = SERIAL_FILL
        cell.font = INPUT_FONT
    elif is_input:
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT
    else:
        cell.font = LOCKED_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_parts_sheet(wb):
    """Main parts inventory sheet — grouped by category."""
    ws = wb.active
    ws.title = "Parts Inventory"

    # Title block
    ws.merge_cells("A1:K1")
    ws["A1"] = "McElroy Fusion Machine Parts Inventory — Underground Solutions"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:K2")
    ws["A2"] = "Fill in BLUE columns (Qty On Hand, Location) and YELLOW columns (Serial Number for tracked assets). Then return this file."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    # Headers at row 4
    headers = [
        ("A", "Category", 22),
        ("B", "Part Name", 38),
        ("C", "McElroy PN", 14),
        ("D", "Description", 50),
        ("E", "Fits Machines", 14),
        ("F", "Tracked /\nSerialized?", 12),
        ("G", "Consumable?", 11),
        ("H", "Qty On Hand\n(FILL IN)", 13),
        ("I", "Location\n(FILL IN)", 20),
        ("J", "Serial Number\n(if tracked)", 18),
        ("K", "Reorder\nPoint", 10),
        ("L", "Reorder\nQty", 9),
        ("M", "Notes\n(FILL IN)", 30),
    ]

    for col_letter, title, width in headers:
        col_idx = ord(col_letter) - ord("A") + 1
        ws.cell(row=4, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width

    style_header_row(ws, 4, len(headers))

    # Location dropdown validation
    loc_list = '"' + ",".join(LOCATIONS) + '"'
    loc_validation = DataValidation(type="list", formula1=loc_list, allow_blank=True)
    loc_validation.error = "Pick a storage location"
    loc_validation.errorTitle = "Invalid Location"
    ws.add_data_validation(loc_validation)

    # Yes/No validation for serialized
    yn_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_validation)

    # Group parts by category
    current_category = None
    row = 5

    # Sort parts by category then name
    sorted_parts = sorted(PARTS, key=lambda p: (p["category"], p["name"]))

    for part in sorted_parts:
        cat = part["category"]

        # Category separator row
        if cat != current_category:
            current_category = cat
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=f"  {cat}")
            cell.font = SUBTITLE_FONT
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER
            for c in range(2, len(headers) + 1):
                ws.cell(row=row, column=c).fill = SECTION_FILL
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        is_serialized = part.get("is_serialized", False)
        is_consumable = part.get("is_consumable", False)

        # Write pre-filled (locked) columns
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=part["name"])
        ws.cell(row=row, column=3, value=part.get("mcelroy_pn", ""))
        ws.cell(row=row, column=4, value=part.get("description", ""))
        ws.cell(row=row, column=5, value=", ".join(part.get("machines", [])))
        ws.cell(row=row, column=6, value="Yes" if is_serialized else "No")
        ws.cell(row=row, column=7, value="Yes" if is_consumable else "No")
        ws.cell(row=row, column=11, value=part.get("reorder_point", 0))
        ws.cell(row=row, column=12, value=part.get("reorder_qty", 0))

        # Style all cells
        for col in range(1, len(headers) + 1):
            is_input = col in (8, 9, 13)  # Qty, Location, Notes
            is_serial_col = col == 10 and is_serialized
            style_data_cell(ws, row, col, is_input=is_input, is_serial=is_serial_col)

        # Add validation to location column
        loc_validation.add(ws.cell(row=row, column=9))

        row += 1

    # Freeze panes
    ws.freeze_panes = "A5"

    # Auto-filter
    ws.auto_filter.ref = f"A4:M{row - 1}"

    return ws


def build_machines_sheet(wb):
    """Fleet sheet — one row per machine, fill in serial numbers."""
    ws = wb.create_sheet("Fleet Machines")

    ws.merge_cells("A1:H1")
    ws["A1"] = "McElroy Fusion Machine Fleet — Underground Solutions"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"] = "Fill in YELLOW columns: real serial numbers, engine hours, and current location."
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)

    headers = [
        ("A", "Machine ID", 16),
        ("B", "Model", 18),
        ("C", "Variant", 10),
        ("D", "Pipe Range", 28),
        ("E", "Power", 32),
        ("F", "McElroy Serial #\n(FILL IN)", 20),
        ("G", "Engine Hours\n(FILL IN)", 14),
        ("H", "Current Location\n(FILL IN)", 22),
        ("I", "Condition\n(FILL IN)", 16),
        ("J", "Notes\n(FILL IN)", 30),
    ]

    for col_letter, title, width in headers:
        col_idx = ord(col_letter) - ord("A") + 1
        ws.cell(row=4, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width

    style_header_row(ws, 4, len(headers))

    # Condition dropdown
    cond_validation = DataValidation(
        type="list",
        formula1='"Good,Fair,Needs Repair,Out of Service,In Shop"',
        allow_blank=True,
    )
    ws.add_data_validation(cond_validation)

    row = 5
    for family, config in FLEET_TEMPLATE.items():
        model_info = next(
            (m for m in MACHINE_MODELS if m["family"] == family and "TracStar" in m["model"]),
            None,
        )
        if not model_info:
            continue

        # Family separator
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = ws.cell(row=row, column=1, value=f"  {family} Family — {model_info['pipe_range']}")
        cell.font = SUBTITLE_FONT
        cell.fill = SECTION_FILL
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        prefix = config["prefix"]

        # Standard machines
        for i in range(1, config["count_standard"] + 1):
            machine_id = f"{prefix}-{i:03d}"
            ws.cell(row=row, column=1, value=machine_id)
            ws.cell(row=row, column=2, value=model_info["model"])
            ws.cell(row=row, column=3, value="Standard")
            ws.cell(row=row, column=4, value=model_info["pipe_range"])
            ws.cell(row=row, column=5, value=model_info["power"])

            for col in range(1, len(headers) + 1):
                is_input = col in (6, 7, 8, 9, 10)
                style_data_cell(ws, row, col, is_input=False, is_serial=(col == 6))
                if col <= 5:
                    ws.cell(row=row, column=col).font = LOCKED_FONT
                if col in (7, 8, 9, 10):
                    ws.cell(row=row, column=col).fill = INPUT_FILL

            cond_validation.add(ws.cell(row=row, column=9))
            row += 1

        # i-series machines
        if model_info.get("model_i"):
            for i in range(1, config["count_i_series"] + 1):
                machine_id = f"{prefix}i-{i:03d}"
                ws.cell(row=row, column=1, value=machine_id)
                ws.cell(row=row, column=2, value=model_info["model_i"])
                ws.cell(row=row, column=3, value="i-series")
                ws.cell(row=row, column=4, value=model_info["pipe_range"])
                ws.cell(row=row, column=5, value=model_info["power"])

                for col in range(1, len(headers) + 1):
                    is_input = col in (6, 7, 8, 9, 10)
                    style_data_cell(ws, row, col, is_input=False, is_serial=(col == 6))
                    if col <= 5:
                        ws.cell(row=row, column=col).font = LOCKED_FONT
                    if col in (7, 8, 9, 10):
                        ws.cell(row=row, column=col).fill = INPUT_FILL

                cond_validation.add(ws.cell(row=row, column=9))
                row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{row - 1}"


def build_locations_sheet(wb):
    """Stock locations reference sheet."""
    ws = wb.create_sheet("Stock Locations")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Stock Locations"
    ws["A1"].font = TITLE_FONT

    headers = [("A", "Location", 28), ("B", "Parent", 20), ("C", "Description", 50)]
    for col_letter, title, width in headers:
        col_idx = ord(col_letter) - ord("A") + 1
        ws.cell(row=3, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width
    style_header_row(ws, 3, len(headers))

    for i, loc in enumerate(STOCK_LOCATIONS, start=4):
        ws.cell(row=i, column=1, value=loc["name"])
        ws.cell(row=i, column=2, value=loc.get("parent", ""))
        ws.cell(row=i, column=3, value=loc.get("description", ""))
        for col in range(1, 4):
            style_data_cell(ws, i, col)


def build_instructions_sheet(wb):
    """How-to sheet."""
    ws = wb.create_sheet("Instructions")

    instructions = [
        ("McElroy Parts Inventory — How to Fill This In", TITLE_FONT),
        ("", None),
        ("This workbook has your full McElroy fleet parts list pre-loaded.", None),
        ("You need to fill in the colored columns and send it back.", None),
        ("", None),
        ("PARTS INVENTORY TAB (blue + yellow columns):", SUBTITLE_FONT),
        ("  - Qty On Hand (blue): How many of this part do you currently have in stock?", None),
        ("  - Location (blue): Where is it stored? Pick from the dropdown.", None),
        ("  - Serial Number (yellow): For tracked assets (facers, heaters, DataLoggers,", None),
        ("    insert sets, cylinders, testers) — enter the McElroy serial number.", None),
        ("  - Notes: Anything relevant — condition, on order, backordered, etc.", None),
        ("", None),
        ("FLEET MACHINES TAB (yellow + blue columns):", SUBTITLE_FONT),
        ("  - McElroy Serial # (yellow): The serial number on the machine nameplate.", None),
        ("  - Engine Hours (blue): Current hour meter reading.", None),
        ("  - Current Location (blue): Where is this machine right now?", None),
        ("  - Condition (blue): Good / Fair / Needs Repair / Out of Service / In Shop", None),
        ("  - Notes: Any issues, upcoming service, etc.", None),
        ("", None),
        ("TIPS:", SUBTITLE_FONT),
        ("  - If you have multiple units of a serialized part, add rows.", None),
        ("  - If a part isn't listed, add it at the bottom of the right category.", None),
        ("  - Don't delete or rename the pre-filled columns.", None),
        ("  - The gray text is reference info — you don't need to change it.", None),
        ("", None),
        ("AFTER FILLING IN:", SUBTITLE_FONT),
        ("  - Save the file and email it back.", None),
        ("  - This same file can be imported into InvenTree if we set that up later.", None),
        ("", None),
        ("Parts data sourced from McElroy operator manuals and Parts Finder.", None),
        ("Part numbers are best-effort — verify before ordering.", None),
    ]

    ws.column_dimensions["A"].width = 80

    for i, (text, font) in enumerate(instructions, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        else:
            cell.font = Font(name="Calibri", size=10)


def main():
    wb = Workbook()

    build_parts_sheet(wb)
    build_machines_sheet(wb)
    build_locations_sheet(wb)
    build_instructions_sheet(wb)

    # Move Instructions to first position
    wb.move_sheet("Instructions", offset=-3)

    out_path = "/home/voidstr3m33/mcelroy-inventory/UGSI_McElroy_Parts_Inventory.xlsx"
    wb.save(out_path)

    # Stats
    n_parts = len(PARTS)
    n_machines = sum(c["count_standard"] + c["count_i_series"] for c in FLEET_TEMPLATE.values())
    n_categories = len(set(p["category"] for p in PARTS))
    n_serialized = sum(1 for p in PARTS if p.get("is_serialized"))
    n_consumable = sum(1 for p in PARTS if p.get("is_consumable"))

    print(f"Workbook saved: {out_path}")
    print(f"  Sheets: Instructions, Parts Inventory, Fleet Machines, Stock Locations")
    print(f"  Parts: {n_parts} ({n_serialized} serialized, {n_consumable} consumable)")
    print(f"  Categories: {n_categories}")
    print(f"  Machines: {n_machines}")
    print(f"  Locations: {len(STOCK_LOCATIONS)}")


if __name__ == "__main__":
    main()
