#!/usr/bin/env python3
"""
load_stock.py  -  Turn the filled inventory workbook into live InvenTree stock.

WHAT IT DOES
  Reads the Parts tab of the filled workbook and, for every row that has an
  On-hand Qty, it:
    1. finds the part in InvenTree by its SKU (IPN),
    2. creates a stock item at the matching location with that quantity,
    3. (optional) sets the part's reorder point from the Reorder Pt column.

USAGE  (run on the machine where InvenTree is running)
    pip install openpyxl requests
    python3 load_stock.py "UGSI_McElroy_Parts_Inventory.xlsx"            # do it
    python3 load_stock.py "UGSI_McElroy_Parts_Inventory.xlsx" --dry-run  # preview only

  Override connection if needed:
    python3 load_stock.py FILE.xlsx --url http://localhost:8080 --user shopmanager --password ugsi2026!

NOTES
  - Safe to re-run: it skips a part if a stock item already exists at that location.
  - Rows with blank/zero qty are ignored (so a half-filled sheet is fine).
  - A small delay between writes keeps the shop-laptop DB from choking like the bulk import did.
"""
import sys, time, argparse
import requests
from openpyxl import load_workbook

def col_map(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v: headers[str(v).strip().lower()] = c
    def find(*names):
        for n in names:
            if n.lower() in headers: return headers[n.lower()]
        return None
    return {
        "sku":   find("SKU (IPN)", "IPN", "SKU", "Part Number (McElroy P/N)"),
        "qty":   find("On-hand Qty", "Qty", "Quantity"),
        "loc":   find("Bin / Location", "Bin", "Location"),
        "reorder": find("Reorder Pt", "Reorder Point"),
    }

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--url", default="http://localhost:8080")
    ap.add_argument("--user", default="shopmanager")
    ap.add_argument("--password", default="ugsi2026!")
    ap.add_argument("--sheet", default="Parts")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--delay", type=float, default=0.15)
    a = ap.parse_args()
    base = a.url.rstrip("/")

    # auth — use basic auth (token endpoint has issues with special chars in password)
    S = requests.Session()
    S.auth = (a.user, a.password)
    S.headers["Content-Type"] = "application/json"

    # verify auth
    r = S.get(f"{base}/api/user/me/", timeout=30)
    if r.status_code != 200:
        sys.exit(f"Auth failed: {r.status_code} {r.text[:200]}")
    print(f"Authenticated as {a.user}")

    # cache locations by lowercase name
    locs = {}
    nxt = f"{base}/api/stock/location/?limit=500"
    while nxt:
        d = S.get(nxt, timeout=30).json()
        for L in d.get("results", d if isinstance(d, list) else []):
            locs[(L["name"] or "").strip().lower()] = L["pk"]
        nxt = d.get("next") if isinstance(d, dict) else None
    print(f"InvenTree has {len(locs)} stock locations")

    wb = load_workbook(a.workbook, data_only=True)
    ws = wb[a.sheet] if a.sheet in wb.sheetnames else wb.active
    cm = col_map(ws)
    if not cm["sku"] or not cm["qty"]:
        sys.exit(f"Could not find SKU and On-hand Qty columns in '{ws.title}'. Found: {cm}")

    created = skipped = nopart = noqty = 0
    for ri in range(2, ws.max_row + 1):
        sku = ws.cell(ri, cm["sku"]).value
        qty = ws.cell(ri, cm["qty"]).value if cm["qty"] else None
        if not sku or sku == "VERIFY": continue
        try: qty = float(qty)
        except (TypeError, ValueError): noqty += 1; continue
        if qty <= 0: noqty += 1; continue
        sku = str(sku).strip()

        # find part by IPN
        pr = S.get(f"{base}/api/part/", params={"IPN": sku}, timeout=30).json()
        results = pr.get("results", pr if isinstance(pr, list) else [])
        if not results:
            print(f"  [no part] {sku}"); nopart += 1; continue
        pid = results[0]["pk"]

        # resolve location (fall back to no location)
        loc_id = None
        if cm["loc"]:
            lname = ws.cell(ri, cm["loc"]).value
            if lname: loc_id = locs.get(str(lname).strip().lower())

        # skip if stock already exists for this part+location
        params = {"part": pid}
        if loc_id: params["location"] = loc_id
        ex = S.get(f"{base}/api/stock/", params=params, timeout=30).json()
        if ex.get("count", len(ex.get("results", []))) > 0:
            skipped += 1; continue

        if a.dry_run:
            print(f"  [would add] {sku}  qty={qty}  loc={loc_id}")
        else:
            payload = {"part": pid, "quantity": qty}
            if loc_id: payload["location"] = loc_id
            sr = S.post(f"{base}/api/stock/", json=payload, timeout=30)
            if sr.status_code in (200, 201):
                # set reorder point if provided
                if cm["reorder"]:
                    rp = ws.cell(ri, cm["reorder"]).value
                    try:
                        if float(rp) > 0:
                            S.patch(f"{base}/api/part/{pid}/",
                                    json={"minimum_stock": float(rp)}, timeout=30)
                    except (TypeError, ValueError): pass
                print(f"  [added] {sku}  qty={qty}")
            else:
                print(f"  [FAIL] {sku}: {sr.status_code} {sr.text[:120]}")
                nopart += 1; continue
        created += 1
        time.sleep(a.delay)

    print("\n=== summary ===")
    print(f"  stock items {'to add' if a.dry_run else 'added'}: {created}")
    print(f"  skipped (already had stock): {skipped}")
    print(f"  rows with no/zero qty (ignored): {noqty}")
    print(f"  SKUs not found in InvenTree: {nopart}")

if __name__ == "__main__":
    main()
